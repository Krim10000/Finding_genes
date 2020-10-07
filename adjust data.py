#adjusting the data
import pandas as pd

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#create excel type item
wb = Workbook()
# select the active worksheet
ws = wb.active

import pickle

#import pickle
#with open("Sequences.txt", "wb") as fp: # Pickling
#    pickle.dump(list, fp)


with open("Sequences.txt", "rb") as fp:   # Unpickling
    
    seq = pickle.load(fp)

df5=pd.DataFrame ({"cadenas":seq})

#Data={"cadenas":["ATG","CTA","TAT"]}
#df3= pd.DataFrame (Data)

for x in range(len(df5.index)): # numero de filas 
    for y in range(len(df5.cadenas[x])): # caracter de la cadena/ columna

        text=(df5.cadenas[x])[y]

        mau=y+1
        column_letter = get_column_letter(mau)
        
        #ws.cell(row=x,column=y)=text
        p=column_letter + str(x+2)
        p="".join(p)
           
        ws[p] = str(text)

for y in range(len(df5.cadenas[1])):
    text=y
    mau=y+1
    column_letter = get_column_letter(mau)

      
    p= column_letter+ str(1)
    p="".join(p)
           
    ws[p] = str(text)
    
    

wb.save("sample.xlsx")


df5=pd.read_excel('sample.xlsx')
df5.insert(0, 'Label', '1')

print("Label 1")
print(df5.head())
print("")




##################



#adjusting the data
import pandas as pd

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#create excel type item
wb = Workbook()
# select the active worksheet
ws = wb.active

import pickle

#import pickle
#with open("Sequences.txt", "wb") as fp: # Pickling
#    pickle.dump(list, fp)


with open("Sequences2.txt", "rb") as fp:   # Unpickling
    
    seq = pickle.load(fp)

df6=pd.DataFrame ({"cadenas":seq})

#Data={"cadenas":["ATG","CTA","TAT"]}
#df3= pd.DataFrame (Data)

for x in range(len(df6.index)): # numero de filas 
    for y in range(len(df6.cadenas[x])): # caracter de la cadena/ columna

        text=(df6.cadenas[x])[y]

        mau=y+1
        column_letter = get_column_letter(mau)
        
        #ws.cell(row=x,column=y)=text
        p=column_letter + str(x+2)
        p="".join(p)
           
        ws[p] = str(text)

for y in range(len(df6.cadenas[1])):
    text=y
    mau=y+1
    column_letter = get_column_letter(mau)

      
    p= column_letter+ str(1)
    p="".join(p)
           
    ws[p] = str(text)
    
    

wb.save("sample0.xlsx")


df6=pd.read_excel('sample0.xlsx')

df6.insert(0, 'Label', '0')
print("Label 0")
print(df6.head())
print("")



dff=pd.concat([df5,df6])
print("Label 0 & 1")

print(dff.head())
dff=dff.reset_index(drop=True)
dff.to_excel("Labels.xlsx")








