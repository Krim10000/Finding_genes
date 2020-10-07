
#This is just to show from where did I obtein the positions of the genes in anotations.xlsx


import re
import pandas as pd
import openpyxl
from openpyxl import Workbook
import datetime 

start_time = datetime.datetime.now()

#filename of the excel file
name ='Resumen.xlsx'

#Name of the file with the DNA info
filename ="NC_000913.3[1..4641652].fa"

wb = Workbook()
ws =  wb.active

#Hoja del archivo
ws5 = wb.create_sheet("Resumen")

#las secuencia de ADN tienen dos posibles sentidos,
#5' a 3', hebra antisentido
#3' a 5'hebra sentido o  coding strand
#La secuencia de ADN en el archivo es 5' a 3' hebra antisentido

f=open(filename, "r")

contents =f.read()
#remueve espacios vacios y nuevas lineas.
contents = re.sub(r"[\n\t\s]*", "", contents)
contents1=contents

#now im going to make 3'5' chain

reverse = contents[::-1]

#Converting T-A, A-T, G-C, C-G

reverse= reverse.replace("A","t").replace("T","A").replace("G","c").replace("C","G").replace("c","C").replace("t","T")

contents = contents+ ("X"*1000)+reverse

#primeros n elementos
n= 10000*20
contents = contents[0:n]#comentar, esta linea es solo para hacelar las pruebas

###########################
def posciones(secuencia):

    minicio = re.finditer(secuencia ,contents)
    pinicio = [match.start() for match in minicio]
    return pinicio

################################
#formato con puntos

def itanum(x):
    return format(x,',d').replace(",",".")
################################

def contar(seq):
    return itanum(len(re.findall(seq,contents)))






#Escherichia coli, emplea en un 83% de los casos ATG (AUG en el ARN), GTG en un 14% (GUG en el transcrito) y en un 3% TTG (UUG en el ARN)
seqI = "ATG|GTG|TTG"#5'3'

#Existen tres codones de terminación, que reciben distintos nombres. «UAG» «codón ámbar»;
#«UGA», como «codón ópalo»;
#y «UAA», como «codón ocre»


seqT = "TAA|TGA|TAG"#5'3'

print("")
print("Numero de bases " +str(itanum(len(contents1))))
print("")
print("Se encontraron "+ str(contar(seqI)) + " repeticiones de la secuencias de inicio " +  seqI)
print("")
print("Se encontraron "+ str(contar(seqT)) + " repeticiones de la secuencias de termino " +  seqT)
print("")

#crear listado de posiciones:

LseqI=posciones(seqI)
LseqT=posciones(seqT)



print("Creando listado de posibles genes  (esta es la parte más lenta)")

##Buscando inicios de genes

#1)Crea una lista con todos los codones de inicio, con un largo de 50 caracateres en total, siendo los 3 ultimos caracteres el codon de inicio respectivo

#2)importa la lista como base de datos en pandas

#3)aplica  en pd.get_dummies

#4)aplica la NN a la base de datos

#5) si la etiqueta es 1 corresponde al inicio de un gen y la posicion pasa a la lista LseqIs

#######1)
LseqI0=[]
#Dummy seqs
LseqI0.append("G"*50)
LseqI0.append("T"*50)
LseqI0.append("A"*50)
LseqI0.append("C"*50)

##############################

size=50


##############################
for x in range(len(LseqI)):
    mauu= (contents[LseqI[x]-(size-3):LseqI[x]+3])
    sumando =mauu.count("A")+mauu.count("T")+mauu.count("G")+mauu.count("C")
    if sumando ==size:
        LseqI0.append(mauu)


#LseqI corresponde a una posicion

#LseqI0 corresponde a una cadena de 50 caracteres

#LseqIz corresponde a los valores en un array que tiene la etiqueta correcta

#LseqIs corresponde a la seleccion de LseqI

#######2)
df=pd.DataFrame(LseqI0)

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb0 = Workbook()
ws0 =  wb0.active

df.columns = ['cadenas']

for x in range(len(df.index)): # numero de filas 
    for y in range(len(df.cadenas[x])): # caracter de la cadena/ columna

        text=(df.cadenas[x])[y]

        mau=y+1
        column_letter = get_column_letter(mau)
        
        #ws.cell(row=x,column=y)=text
        p=column_letter + str(x+2)
        p="".join(p)
           
        ws0[p] = str(text)
        if x%10000==0 and y==0:
            print(str(itanum(x))+"/"+str(itanum(len(df.index))))
            
print("")
for y in range(len(df.cadenas[1])):
    text=y
    mau=y+1
    column_letter = get_column_letter(mau)

      
    p= column_letter+ str(1)
    p="".join(p)
           
    ws0[p] = str(text)
    
    

wb0.save("DATA.xlsx")
print("Datos de codones de inicio exportados")
print("")
#######3)
df2=pd.read_excel('DATA.xlsx')

df2= pd.get_dummies(df2)#ACGT
print("Datos de codones de inicio importados y formato ajustado")
print("")



#######4)


import pickle

with open("NN1mlp.txt", "rb") as fp:   # Unpickling    
    mlp = pickle.load(fp)

print("NN importada")
print("")

LseqIz=[]
label= mlp.predict(df2)
i=0
for y in range(4,len(df2.index)):
    if label[y] ==1:
        i=i+1
        LseqIz.append(y-4)#because I added 4 more A*50,C*50,T*50, G*50 to have a correct data shape




#convertir los valores correctos en las posiciones correspondientes
LseqIs=[]

for x in range(len(LseqIz)):
    LseqIs.append(LseqI[int(LseqIz[x])])

print("se han selesccionado " + str(len(LseqIz))+ "/" +str(len(LseqI)-4))
porcentaje =len(LseqIz)/(len(LseqI)-4)*100
print(str(porcentaje) +" %")


        
        
print("")
print(str(itanum(len(LseqIs))) + " Secuencias de inicio identificas como correctas")
print("")            

wb.save(filename = name)

#print("Numero de promotores y secuencias de inicio con distancia correcta " + str(itanum(len(LseqIs))))

print("")

print("Creando listado de posibles genes")
##################################################################

#LseqT corresponde a una posicion

#LseqT0 corresponde a una cadena de 50 caracteres

#LseqTz corresponde a los valores en un array que tiene la etiqueta correcta LseqTz<LseqT0

#LseqTs corresponde a la seleccion de LseqT


#Buscando finales

##seqT = "TAA|TGA|TAG"#5'3'



size=50
LseqT0=[]
LseqT0.append("G"*50)
LseqT0.append("T"*50)
LseqT0.append("A"*50)
LseqT0.append("C"*50)

##############################
for x in range(len(LseqT)):
    mauu= (contents[LseqT[x]:LseqT[x]+size])
    
    sumando =mauu.count("A")+mauu.count("T")+mauu.count("G")+mauu.count("C")
    if sumando ==size:
        LseqT0.append(mauu)


#LseqT0 cadenas de 50 caracteres, len(LseqT0)<len(LseqT)

#######2)
df3=pd.DataFrame(LseqT0)

#>>> df3.index
#RangeIndex(start=0, stop=1861, step=1) #correcto 1857+4

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb1 = Workbook()
ws1 =  wb1.active

df3.columns = ['cadenas']

for x in range(len(df3.index)): # numero de filas 
    for y in range(len(df3.cadenas[x])): # caracter de la cadena/ columna

        text=(df3.cadenas[x])[y]

        #if y==0:
         #   print(x)

        mau=y+1
        column_letter = get_column_letter(mau)
        
        #ws.cell(row=x,column=y)=text
        p=column_letter + str(x+2)
        p="".join(p)
           
        ws1[p] = str(text)
        if x%10000==0 and y==0:
            print(str(itanum(x))+"/"+str(itanum(len(df3.index))))
            
print("")
for y in range(len(df3.cadenas[1])):
    text=y
    mau=y+1
    column_letter = get_column_letter(mau)

      
    p= column_letter+ str(1)
    p="".join(p)
           
    ws1[p] = str(text)

    


    
wb1.save("DATA_end.xlsx")
print("Datos codones de termino exportados")
print("")
#######3)
df4=pd.read_excel('DATA_end.xlsx')
print(len(df4.index))


df5= pd.get_dummies(df4)#ACGT
print("Datos  codones de termino importados y formato ajustado")
print("")



#######4)


import pickle

with open("NN_end_mlp.txt", "rb") as fp:   # Unpickling    
    mlp2 = pickle.load(fp)

print("NN 2 importada")
print("")

LseqTz=[]
label2= mlp2.predict(df5)
i=0
for y in range(4,len(df5.index)):
    if label2[y] ==2:
        i=i+1
        LseqTz.append(y-4)#because I added 4 more A*50,C*50,T*50, G*50 to have a correct data shape

print("se han selesccionado " + str(len(LseqTz))+ "/" +str(len(LseqT)-4))
porcentaje =len(LseqTz)/(len(LseqT)-4)*100
print(str(porcentaje) +" %")


#convertir los valores correctos en las posiciones correspondientes
LseqTs=[]

for x in range(len(LseqTz)):
    LseqTs.append(LseqT[int(LseqTz[x])])

#(LseqTz) casos encontrados por MLP
#LseqT casos totales 

    



#################################################################

#Genes con inicio y final
CINICIO=[]
ADN =[]
GEN =[]
GENf=[]
k=1
G=0


########################### RESCRIBIR
#BUSCAR GENES ENTRE ZONAS DE INICIO TERMINO
#LseqIs   INICIO OPERON
#LseqTs   TERMINO OPERON

for x in range(len(LseqIs)):
    for y in range(len(LseqTs)):


        if G==len(LseqIs):#cuando termina de recorrer rompe el ciclo
            break

        if ((int(LseqTs[y])>int(LseqIs[x])) and (int(LseqTs[y])-int(LseqIs[x]))%3==0):                          
            if (len(contents[LseqIs[x]+3:LseqTs[y]])>20): #and (LseqIs[x]-Lp2ss[x])>5:
                #cantidad minima de aminoacidos 8(7*3=21) and  #el promotor no se sobrepone al codon de inicio         

                k=k+1
                G=G+1#Contador numero de genes
                
##                ws2.cell(row=int(k), column=2).value = inicio_gen[x]
##                ws2.cell(row=int(k), column=3).value = LseqT[y]
                
                GENf.append(LseqTs[y])
                
##                ws2.cell(row=int(k), column=4).value = contents[inicio_gen[x]+3:LseqT[y]]#gen sin inicio ni termino
                
                GEN.append(contents[LseqIs[x]+3:LseqTs[y]])
                ADN.append(contents[LseqIs[x]+3:LseqTs[y]])
                
                CINICIO.append(contents[LseqIs[x]:LseqIs[x]+3])
                
##                ws2.cell(row=int(k), column=5).value = contents[inicio_gen[x]:LseqT[y]+3]#gen con inicio y termino
##                ws2.cell(row=int(k), column=6).value = LseqT[y]-inicio_gen[x]
                

                #resumen

                #ws5.cell(row=int(k), column=1).value = contents[Lp1ss[x]:Lp1ss[x]+len(promotor_1)]#promotor_1
                #ws5.cell(row=int(k), column=2).value = Lp1ss[x]#inicio1
                #ws5.cell(row=int(k), column=3).value= Lp1ss[x]-LseqIs[x]# distancia
                #ws5.cell(row=int(k), column=4).value = contents[Lp2ss[x]:Lp2ss[x]+len(promotor_2)]#promotor_2
                #ws5.cell(row=int(k), column=5).value = Lp2ss[x]#inicio2
                #ws5.cell(row=int(k), column=6).value= Lp2ss[x]-LseqIs[x]
                ws5.cell(row=int(k), column=7).value = contents[LseqIs[x]:LseqIs[x]+3]#codon inicio
                ws5.cell(row=int(k), column=8).value = contents[LseqIs[x]+3:LseqTs[y]] #secuencia codificante 5'3'
                ws5.cell(row=int(k), column=9).value = contents[LseqT[y]:LseqTs[y]+3]#codon termino
                ws5.cell(row=int(k), column=10).value= LseqIs[x]#inicio gen
                ws5.cell(row=int(k), column=11).value= (LseqTs[y]+3)#fin gen

                if G%500==0:
                    print("Se ha encontrado el codon de termino de "+ str(G)+ " genes")
                    print("")


                    
                #Cuando identifica el fin de un gen, pasa al siguiente
                break

            else:
                break
##################################################################
print(str(G) +" numero de genes con tamaño apropiado")

#mRna

print("")

print("Creando mRna")

#Convierte el DNA en mRNA Transcripción
mRNA=[]
mRNA= GEN
for x in range (len(mRNA)):
    #mRNA[x]=mRNA[x].replace("A","U").replace("T","A").replace("G","c").replace("C","G").replace("c","C")#3'5'
    mRNA[x]=mRNA[x].replace("T","U")#5'3'
    
    


##################################################################
#Escritura hoja 3
##ws3 = wb.create_sheet("mRNA")
##ws3.cell(row=1, column=2).value="inicio gen"
##ws3.cell(row=1, column=3).value="termino gen"
##ws3.cell(row=1, column=4).value="mRna"
##ws3.cell(row=1, column=5).value="largo"
##ws3.cell(row=1, column=6).value="Codon inicio"




k=1
for x in range (len(mRNA)):
    k=k+1#Contador para las filas
##    ws3.cell(row=k, column=2).value=inicio_gen[x]
##    ws3.cell(row=k, column=3).value=GENf[x]
##    ws3.cell(row=k, column=4).value=mRNA[x]
##    ws3.cell(row=k, column=5).value=len(mRNA[x])
##    ws3.cell(row=k, column=6).value=CINICIO[x]
    #resumen
    ws5.cell(row=k, column=12).value= mRNA[x]
##################################################################
wb.save(filename = name)

print(str(len(mRNA)) + " numero de mRNA registrados")

print("")

print("Creando cadenas de AA")




#Convierte el mRNA en AA Traducción
AA=[]


for x in range(len(mRNA)):
    y= len(mRNA[x])
    pep = "M"
    z=0
    while z <y/3:
        ##########
        pep0=""
        codon = mRNA[x][z*3:z*3+3]
        #print(codon)
        #85841

        #casos, si codon es reemplzar por
        #if re.search(r'pattern', string):
        if re.search(r'GC.',codon):#1 BIEN
            pep0="A"
        elif re.search(r'CG.|AGA|AGG',codon):#2 BIEN
            pep0="R"
        elif re.search(r'AAU|AAC',codon):#3 BIEN
            pep0="N"
        elif re.search(r'GAU|GAC',codon):#4 BIEN
            pep0="D"
        elif re.search(r'UAA|UGA|UAG', codon):#6 #stop BIEN
            pep0="."
        elif re.search(r'UGU|UGC', codon):#7 BIEN
            pep0="C"
        elif re.search(r'CAA|CAG', codon):#8 BIEN
            pep0="Q"
        elif re.search(r'GAA|GAG', codon):#9 BIEN
            pep0="E"
        elif re.search(r'CC.', codon):#18 BIEN
            pep0="P"
        elif re.search(r'UC.|AGU|AGC', codon):#19 BIEN
            pep0="S"
        elif re.search(r'AC.',codon):#20  
            pep0="T"
        elif re.search(r'UGG',codon):#21
            pep0="W"
        elif re.search(r"UAU|UAC",codon):#22 
            pep0="Y"
        elif re.search(r'GU.',codon):#23
            pep0="V"
        elif re.search(r'GG.', codon):#11 BIEN
            pep0="G"
        elif re.search(r'CAU|CAC', codon):#12 BIEN
            pep0="H"
        elif re.search(r'AUU|AUC|AUA',codon):#13 BIEN
            pep0="I"
        elif  re.search(r'CU.|UUA|UUG',codon):#14 BIEN 
            pep0="L"
        elif re.search(r'AAA|AAG', codon):#15 BIEN
            pep0="K"
        elif re.search(r'AUG', codon):#16 BIEN
            pep0="M"
        elif re.search(r'UUU|UUC', codon):#17 BIEN
            pep0="F"
        else: # literaly something else
            pep0="!"




        #crea el peptido
        pep = pep+pep0
        z=z+1
    AA.append(pep)

##################################################################

k=1
for x in range(len(GEN)):
        k=k+1
      
        ws5.cell(row=k, column=13).value=AA[x]
        ws5.cell(row=k, column=14).value=len(AA[x])
        



wb.save(filename = name)

print(str(len(mRNA)) +  " numero de cadenas de AA registradas")

print("")

##################################################################


#cadena mRNA de largo "y" ej 9
#cadena AA de largo "z" ej 3
#los bloques
#0-2 [0:3]       0
#3-5 [3:6]       1
#6-8 [6:9]       2
#    [z*3:z*3+3]     y/3 = 3


##################################################################



#resumen
ws5.cell(row=1, column=1).value="Promotor 1"
ws5.cell(row=1, column=2).value="Inicio promotor 1"
ws5.cell(row=1, column=3).value="Distancia -35"
ws5.cell(row=1, column=4).value="Promotor 2"
ws5.cell(row=1, column=5).value="Inicio promotor 2"
ws5.cell(row=1, column=6).value="Distancia -10"
ws5.cell(row=1, column=7).value="Codon inicio"
ws5.cell(row=1, column=8).value="Gen 5'3'(sin codones de inicio ni termino)"
ws5.cell(row=1, column=9).value="Codon termino"
ws5.cell(row=1, column=10).value="Inicio gen"
ws5.cell(row=1, column=11).value="Fin gen"
ws5.cell(row=1, column=12).value="mRNA"
ws5.cell(row=1, column=13).value="Proteina"
ws5.cell(row=1, column=14).value="Largo proteina"


#############################################################
#CHECK

Protein_filename = "UP000000625_83333.fasta"
G=open(Protein_filename, "r")

Check =G.read()

start = 'SV=\d' # SV= + numero
#https://www.uniprot.org/help/fasta-headers
end = '>sp|>tr'
def posciones(secuencia):

    minicio = re.finditer(secuencia ,Check)
    pinicio = [match.start() for match in minicio]
    return pinicio

Lstart=posciones(start)

Lend=posciones(end)

#Creacion de listas
Lseq=[]
Ldesq=[]
for x in range(len(Lstart)):
    for y in range(len(Lend)):
        if Lend[y]>Lstart[x]:
            Lseq.append(Check[Lstart[x]+4:Lend[y]])
            Ldesq.append(Check[Lend[y-1]:Lstart[x]+4])
            
            break
#ultimo elemento        
Lseq.append(Check[Lstart[-1]+4:len(Check)])
Ldesq.append(Check[Lend[-1]:Lstart[-1]])

for x in range(len(Lseq)):
    #remueve espacios vacios y nuevas lineas.
    Lseq[x] = re.sub(r"[\n\t\s]*", "", Lseq[x])


ws5.cell(row=1, column=15).value="Match"
ws5.cell(row=1, column=16).value="OProtein"
ws5.cell(row=1, column=17).value="Data"
ws5.cell(row=1, column=18).value="%"

#re.search(r'pattern', string)
Lseqs=[]
Ldesqs=[]
Lmatch=[]



#remueve espacios vacios y nuevas lineas.
Check = re.sub(r"[\n\t\s]*", "", Check)
from openpyxl.styles import PatternFill

j=0
k=1

for x in range(len(AA)):
    
    pattern = AA[x]
   
    if re.search(pattern, Check):
        k=k+1
        ws5.cell(row=k, column=15).value=int(1)
        ws5.cell(row=k, column=15).fill =PatternFill(start_color="8a2be2", end_color="8a2be2", fill_type="solid")#relleno morado
        j=j+1
    else:
        k=k+1
        ws5.cell(row=k, column=15).value=int(0)


print("Se encontraron " + str(j) + " proteina correctas")




wb.save(filename = name)


print("")

            


total_time=(datetime.datetime.now()-start_time)

print("tiempo de ejecución " + str(total_time.seconds)+ " segundos")

print("Programa terminado")
